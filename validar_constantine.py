"""
Validador de WhatsApp - Edifício Constantine (Prudente de Moraes, 65)
Usa a API 2chat para checar quais números têm WhatsApp.
Execute: python3 validar_constantine.py
"""

import re
import requests
import time
import json
import openpyxl
from openpyxl.styles import PatternFill, Font

API_KEY = "UAK3057d948-badf-43be-ada2-d94f2ac80814"
MEU_NUMERO = "+5544991529335"
BASE_URL = "https://api.p.2chat.io/open/whatsapp/check-number"

PROPRIETARIOS = [
    {"apt": "44",   "owner": "CARLOS AFONSO MARQUES",                    "numbers": ["+5544999723481", "+5544999723483"]},
    {"apt": "101",  "owner": "JURANDIR FERNANDO COMAR",                   "numbers": ["+5542999053336"]},
    {"apt": "102",  "owner": "ANDERSON ANDRE BIONDO",                     "numbers": ["+5544999731590", "+5544910054586"]},
    {"apt": "201",  "owner": "JOSE APARECIDO CALIXTO ROSA",               "numbers": ["+5544999703018", "+5544991078706"]},
    {"apt": "202",  "owner": "MARCIO GUSHIKEN",                           "numbers": ["+5544999418887", "+5544991449979"]},
    {"apt": "302",  "owner": "IDALGO ANTONIOLLI",                         "numbers": ["+5544991666320", "+5544991671110", "+5544991661591"]},
    {"apt": "503",  "owner": "FERNANDA OKAMOTO TOSCANO",                  "numbers": ["+5511982099902", "+5511964688864"]},
    {"apt": "504",  "owner": "ALTEMIR SERGIO MOLINA",                     "numbers": ["+5544991090818", "+5544999105888", "+5544999385888"]},
    {"apt": "804",  "owner": "EMERSON VIERA",                             "numbers": ["+5543988139457", "+5543999761999"]},
    {"apt": "901",  "owner": "LUDMILA PRISCILLA MANETTI",                 "numbers": ["+5566999853662", "+5544998171111"]},
    {"apt": "1001", "owner": "JOSE SUSUMU SAKAMOTO",                      "numbers": ["+5544999918000", "+5544988480000"]},
    {"apt": "1002", "owner": "MARCELO HENRIQUE DOS SANTOS BORTOLOCCI",    "numbers": ["+5544991260944", "+5544999640092"]},
    {"apt": "1003", "owner": "ALINE BRAGA DRUMMOND",                      "numbers": ["+5544984172062", "+5544998365292", "+5544984531646"]},
    {"apt": "1004", "owner": "CRISTINA GIATTI MARQUES DE SOUZA",          "numbers": ["+5544999161629"]},
    {"apt": "1102", "owner": "CELSO SEITIRO OTAKE",                       "numbers": ["+5544999613958", "+5544991022332"]},
    {"apt": "1103", "owner": "ANDERSON CLAYTON GOMES",                    "numbers": ["+5544999221399", "+5544984849933"]},
    {"apt": "1104", "owner": "FREDERICO AUGUSTO STIPP GROCHAU",           "numbers": ["+5544991135342", "+5544910071004"]},
    {"apt": "1201", "owner": "RODRIGO RIBEIRO SEZINI",                    "numbers": ["+5544991140030", "+5544999736453"]},
    {"apt": "1203", "owner": "JOSE ROBERTO GIATTI",                       "numbers": ["+5544999564100", "+5544998070303"]},
    {"apt": "1204", "owner": "JOAO VITOR GIATTI",                         "numbers": ["+5544998079406", "+5544998070010"]},
    {"apt": "1301", "owner": "OSCAR AUGUSTO MADER",                       "numbers": ["+5544991174040", "+5544991489622"]},
    {"apt": "1302", "owner": "SETSUKO MORIMOTO YOSHIOKA",                 "numbers": ["+5544999914534", "+5544991074534"]},
    {"apt": "1303", "owner": "HIROME TAKASHIMA",                          "numbers": ["+5544991191898", "+5544991174044"]},
    {"apt": "1304", "owner": "CLEVERSON RODRIGUES TALARICO",              "numbers": ["+5544991580503", "+5544991050825"]},
    {"apt": "1401", "owner": "JOAO EUCLIDES MENEZES",                     "numbers": ["+5544999695501", "+5544999690541"]},
    {"apt": "1402", "owner": "WALDEMAR YAMAMOTO",                         "numbers": ["+5544999855030", "+5544998550103"]},
    {"apt": "1403", "owner": "IVANILDO CORREIA DE OLIVEIRA",              "numbers": ["+5544999639060", "+5544999639061"]},
    {"apt": "1404", "owner": "ITAMAR RODRIGUES FILHO",                    "numbers": ["+5544999809977", "+5544999932988"]},
    {"apt": "1501", "owner": "JULIENY CRISTINA PEREIRA LOPES PAULINO",   "numbers": ["+5544998181808", "+5544998181806"]},
    {"apt": "1502", "owner": "ANDREA PARRA MANREZA CARVALHO",             "numbers": ["+5544991482590", "+5544991061040"]},
    {"apt": "1503", "owner": "ALESSANDRA MILANESI TOSI",                  "numbers": ["+5544991188814"]},
    {"apt": "1504", "owner": "SILVANA TIEMY MIURA HOSOKAWA",              "numbers": ["+5544999903484", "+5544998365555"]},
    {"apt": "1601", "owner": "POLYANE ALVES LIMA YOKOTA",                 "numbers": ["+5544999464012", "+5544997088972"]},
    {"apt": "1602", "owner": "MARIA TEREZA RAMOS TRENTINI",               "numbers": ["+5544991501226", "+5544999550049"]},
    {"apt": "1603", "owner": "FLAVIA CARLIN PEREIRA MENDES",              "numbers": ["+5544999012049", "+5544991432049"]},
    {"apt": "1604", "owner": "ELTON HENRIQUE RIBEIRO DE OLIVEIRA",        "numbers": ["+5544991475200", "+5544991275200"]},
    {"apt": "1701", "owner": "MARCELINO MAMORU TAKASHIMA",                "numbers": ["+5544991174040", "+5544991489622"]},
    {"apt": "1702", "owner": "TAQUECO YAMAMOTO",                          "numbers": ["+5544999719971", "+5544998560512"]},
    {"apt": "1703", "owner": "PRISCILA AKEMI MANABE SATO",                "numbers": ["+5544991111434", "+5544998316698"]},
    {"apt": "1704", "owner": "JOSE MARCOS ROMAGNOLLI",                    "numbers": ["+5544991777447", "+5544999067101"]},
    {"apt": "1801", "owner": "PAULO CESAR CASAGRANDE",                    "numbers": ["+5544991113020", "+5544999744884"]},
    {"apt": "1802", "owner": "MARIA LUCIA BATISTA NUNES",                 "numbers": ["+5544998036040", "+5544999634840"]},
    {"apt": "1803", "owner": "WELLINGTON CORADINI",                       "numbers": ["+5544999920066", "+5544984324840"]},
    {"apt": "1804", "owner": "ROGERIO JUNIO PEREIRA",                     "numbers": ["+5544999659044", "+5544999659041"]},
    {"apt": "1901", "owner": "GERALDO AMARO RAMOS FILHO",                 "numbers": ["+5544991475200", "+5544991275200"]},
    {"apt": "1902", "owner": "ORLANDO CARLOS GOMES COLHADO",              "numbers": ["+5544998279146", "+5544991045057"]},
    {"apt": "1903", "owner": "ROSELI PEREIRA SANTOS",                     "numbers": ["+5544998168832", "+5544999201001"]},
    {"apt": "2001", "owner": "OSVALDO MASSAYUKI HARA",                    "numbers": ["+5544991495050", "+5544998344442"]},
    {"apt": "2002", "owner": "PAULO ROBERTO DE OLIVEIRA",                 "numbers": ["+5544999751900", "+5544999751955"]},
    {"apt": "2003", "owner": "VALERIA APARECIDA MARQUES",                 "numbers": ["+5544991728848", "+5544991395558"]},
    {"apt": "2004", "owner": "JOAO MARCOS CASAGRANDE",                    "numbers": ["+5544991113020", "+5544999744884"]},
    {"apt": "2101", "owner": "UMBERTO KATSUMASA TANAKA",                  "numbers": ["+5544999674012", "+5544999640623"]},
    {"apt": "2104", "owner": "KESIA GEMIMA PALMA RIGO WUTZOW",            "numbers": ["+5544998367765", "+5544991722345"]},
    {"apt": "2203", "owner": "PERSIO ACHOA CLAUDINO",                     "numbers": ["+5544991727549", "+5544991583028", "+5544991182733"]},
    {"apt": "2303", "owner": "GABRIEL SAMESIMA BIM",                      "numbers": ["+5544991537335", "+5561999828015"]},
    {"apt": "2401", "owner": "JOAO EDSON CHAVENCO",                       "numbers": ["+5544999848451"]},
]

def check_number(number):
    url = f"{BASE_URL}/{MEU_NUMERO}/{number}"
    try:
        resp = requests.get(url, headers={"X-User-API-Key": API_KEY}, timeout=15)
        data = resp.json()
        return data.get("on_whatsapp", False)
    except Exception as e:
        print(f"  ERRO: {e}")
        return None

def main():
    results = []
    total = sum(len(p["numbers"]) for p in PROPRIETARIOS)
    count = 0

    print(f"Validando {total} números de {len(PROPRIETARIOS)} proprietários...\n")

    for prop in PROPRIETARIOS:
        number_results = []
        for num in prop["numbers"]:
            count += 1
            on_wa = check_number(num)
            status = "✅ WhatsApp" if on_wa else ("❌ Sem WA" if on_wa is False else "⚠️ Erro")
            print(f"[{count}/{total}] Apto {prop['apt']} | {num} | {status}")
            number_results.append({"number": num, "on_whatsapp": on_wa})
            time.sleep(0.5)

        wa_numbers = [n["number"] for n in number_results if n["on_whatsapp"]]
        results.append({
            "apt": prop["apt"],
            "owner": prop["owner"],
            "has_whatsapp": len(wa_numbers) > 0,
            "wa_numbers": wa_numbers,
            "all_numbers": number_results,
        })

    # Descobrir máximo de números com WA para criar colunas dinâmicas
    max_wa = max((len(r["wa_numbers"]) for r in results), default=1)

    # Gerar Excel — uma linha por proprietário, colunas separadas para cada número com WA
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validação WhatsApp"

    num_cols = [f"WhatsApp {i+1}" for i in range(max_wa)]
    headers = ["Apartamento", "Proprietário", "Tem WhatsApp?"] + num_cols
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFCDD2")
    yellow = PatternFill("solid", fgColor="FFF9C4")

    for r in results:
        wa_cols = r["wa_numbers"] + [""] * (max_wa - len(r["wa_numbers"]))
        row = [r["apt"], r["owner"], "SIM" if r["has_whatsapp"] else "NÃO"] + wa_cols
        ws.append(row)
        fill = green if r["has_whatsapp"] else red
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    for i in range(max_wa):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = 22

    output = "Constantine_WhatsApp_Validado.xlsx"
    wb.save(output)

    total_wa = sum(1 for r in results if r["has_whatsapp"])
    print(f"\n✅ Concluído! {total_wa}/{len(results)} proprietários com WhatsApp.")
    print(f"📊 Planilha salva: {output}")

if __name__ == "__main__":
    main()
