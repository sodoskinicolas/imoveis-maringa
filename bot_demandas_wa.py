#!/usr/bin/env python3
"""
bot_demandas_wa.py
Insere demandas (corretores buscando imóveis) em Demandas_Grupos.xlsx.

Uso:
  python bot_demandas_wa.py --dados '{"grupo":"INNOVARE","corretor":"Ana","contato":"554499...",
                                      "tipo":"Apartamento","regiao":"Zona 7","area_min":"70",
                                      "quartos":"3","orcamento":"420000","obs":"URGENTE | Sacada"}'
  python bot_demandas_wa.py --arquivo demandas.json
"""

import sys, json, argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Erro: pip install openpyxl"); sys.exit(1)

PLANILHA = Path(__file__).parent / "Demandas_Grupos.xlsx"

COL_ORDER = ["data", "grupo", "corretor", "contato",
             "tipo", "regiao", "area_min", "quartos",
             "suites", "vagas", "orcamento", "obs", "status"]

ALT_FILL   = PatternFill("solid", start_color="F0EAFB")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
BORDER     = Border(
    left=Side(style="thin", color="C4B0E8"),
    right=Side(style="thin", color="C4B0E8"),
    top=Side(style="thin", color="C4B0E8"),
    bottom=Side(style="thin", color="C4B0E8"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP   = Alignment(vertical="center", wrap_text=True)


def normalizar(item: dict) -> dict:
    out = {k: "" for k in COL_ORDER}
    out.update({"data": datetime.now().strftime("%Y-%m-%d"), "status": "Novo"})
    mapa = {
        "grupo":    ["grupo", "group"],
        "corretor": ["corretor", "broker", "nome"],
        "contato":  ["contato", "telefone", "whatsapp", "phone"],
        "tipo":     ["tipo", "tipo_buscado", "type"],
        "regiao":   ["regiao", "região", "bairro", "localizacao", "location", "area_interesse"],
        "area_min": ["area_min", "area", "metragem_min", "m2_min"],
        "quartos":  ["quartos", "bedrooms", "dormitorios"],
        "suites":   ["suites", "suítes"],
        "vagas":    ["vagas", "garagem"],
        "orcamento":["orcamento", "orçamento", "preco_max", "valor_max", "budget", "preco"],
        "obs":      ["obs", "observacoes", "observações", "descricao", "description"],
    }
    for campo, chaves in mapa.items():
        for chave in chaves:
            if chave in item and item[chave]:
                out[campo] = str(item[chave]).strip(); break
    return out


def ja_existe(ws, item: dict) -> bool:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        if (str(row[2] or "").lower() == item["corretor"].lower() and
                str(row[5] or "").lower() == item["regiao"].lower() and
                str(row[10] or "") == item["orcamento"]):
            return True
    return False


def inserir(item: dict):
    if not PLANILHA.exists():
        print(f"Planilha não encontrada: {PLANILHA}"); sys.exit(1)

    wb = load_workbook(PLANILHA)
    ws = wb["Demandas"]
    item = normalizar(item)

    if ja_existe(ws, item):
        print(f"⚠️  Duplicata: {item['corretor']} / {item['regiao']} / R${item['orcamento']}")
        return

    next_row = ws.max_row + 1
    fill = ALT_FILL if next_row % 2 == 0 else WHITE_FILL

    for col_idx, val in enumerate([item[k] for k in COL_ORDER], 1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.fill = fill; cell.border = BORDER
        cell.font = Font(name="Arial", size=10)
        cell.alignment = WRAP if col_idx in (6, 12) else CENTER

    wb.save(PLANILHA)
    print(f"✅ Demanda: {item['corretor']} | {item['regiao']} | até R$ {item['orcamento']}")
    _regenerar_site()


def _regenerar_site():
    import subprocess, sys as _sys
    s = Path(__file__).parent / "gerar_site.py"
    if s.exists():
        subprocess.run([_sys.executable, str(s)], check=False)


def processar_arquivo(path):
    with open(path, "r", encoding="utf-8") as f:
        dados = json.load(f)
    for item in (dados if isinstance(dados, list) else [dados]):
        inserir(item)


def main():
    parser = argparse.ArgumentParser()
    g = parser.add_mutually_exclusive_group(required=True)
    g.add_argument("--dados")
    g.add_argument("--arquivo")
    args = parser.parse_args()
    if args.dados: inserir(json.loads(args.dados))
    else: processar_arquivo(args.arquivo)


if __name__ == "__main__":
    main()
