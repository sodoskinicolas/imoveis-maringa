#!/usr/bin/env python3
"""
processar_mensagens.py
Lê mensagens capturadas pelo bot Baileys, agrupa fotos + texto do mesmo corretor
como um único imóvel, extrai dados e atualiza Imoveis_Grupos.xlsx.

Uso:
    python3 processar_mensagens.py             # processa e atualiza planilha
    python3 processar_mensagens.py --dry-run   # mostra sem salvar
    python3 processar_mensagens.py --ver-fila  # lista mensagens pendentes
"""

import json, re, sys, os, base64
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

BASE_DIR     = Path(__file__).parent
FILA_FILE    = BASE_DIR / "mensagens_fila.json"
PLANILHA     = BASE_DIR / "Imoveis_Grupos.xlsx"
ABA_IMOVEIS      = "Imóveis"
ABA_DEMANDAS     = "Demandas"
ABA_CONDOMINIOS  = "Condomínios"

DRY_RUN  = "--dry-run"  in sys.argv
VER_FILA = "--ver-fila" in sys.argv

# Janela de tempo para agrupar fotos + texto do mesmo corretor (segundos)
JANELA_AGRUPAMENTO = 300  # 5 minutos

# ─── Colunas (igual ao existente na planilha) ────────────────────────────────
COLUNAS_IMOVEIS = [
    'Data Captura', 'Grupo', 'Corretor', 'Contato (WhatsApp)', 'Tipo',
    'Bairro / Endereço', 'Área (m²)', 'Quartos', 'Suítes', 'Banheiros',
    'Vagas', 'Preço (R$)', 'Observações', 'Status', 'Data Publicação'
]

COLUNAS_DEMANDAS = [
    'Data', 'Grupo', 'Corretor', 'Contato', 'Tipo Buscado', 'Bairro/Região',
    'Área Mín', 'Quartos', 'Suítes', 'Banheiros', 'Vagas', 'Orçamento Máx',
    'Observações', 'Status'
]

COLUNAS_CONDOMINIOS = [
    'Nome', 'Endereço', 'Bairro', 'CEP', 'Construtora / Incorporadora',
    'Ano Lançamento', 'Previsão Entrega', 'Padrão',
    'Torres', 'Andares', 'Total Aptos',
    'Área Mín (m²)', 'Área Máx (m²)', 'Quartos', 'Vagas',
    'Lazer', 'Faixa de Preço', 'Observações', 'Site / Link', 'Data Cadastro'
]

# Nomes de condomínios descobertos nesta execução (para pesquisar ao final)
_CONDOS_NOVOS: set = set()

# ─── Anthropic API ────────────────────────────────────────────────────────────

def _api_key():
    env = BASE_DIR / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            if line.startswith("ANTHROPIC_API_KEY="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("ANTHROPIC_API_KEY", "")

def analisar_imagem(img_path, caption="", autor=""):
    """Claude Haiku analisa uma imagem e extrai dados do imóvel."""
    api_key = _api_key()
    if not api_key or not Path(img_path).exists():
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        img_b64 = base64.standard_b64encode(Path(img_path).read_bytes()).decode()

        prompt = (
            "Você é especialista em imóveis de Maringá/PR. Analise esta imagem de grupo de corretores.\n"
            "Retorne SOMENTE um JSON válido:\n"
            '{"eh_imovel":true/false,"tipo":"Apartamento|Casa|Terreno|Sala Comercial|Outro",'
            '"bairro":"nome ou null","area":numero_m2_ou_null,"quartos":num_ou_null,'
            '"suites":num_ou_null,"banheiros":num_ou_null,"vagas":num_ou_null,'
            '"preco":inteiro_reais_ou_null,"obs":"info extra"}\n\n'
            f"Legenda: {caption or '(sem legenda)'}\nCorretor: {autor}"
        )
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=500,
            messages=[{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":img_b64}},
                {"type":"text","text":prompt}
            ]}]
        )
        m = re.search(r'\{.*\}', resp.content[0].text, re.DOTALL)
        return json.loads(m.group()) if m else None
    except Exception as e:
        print(f"  ⚠️  Claude API: {e}")
        return None

# ─── Classificação: venda vs demanda ─────────────────────────────────────────

RE_DEMANDA = re.compile(
    r'cliente\s+(?:aprov|busc|quer|prec|comprad|procur)|'
    r'tenho\s+cliente|tenho\s+comprador|'
    r'\bpreciso\s+de\b|\bprocuro\b|\bestou\s+procurando\b|\bà\s+procura\b|'
    r'quero\s+(?:comprar|alugar)|'
    r'comprador\s+(?:busca|procura|quer|aprov)|'
    r'aprovado\s+em|aprovada\s+em|financiamento\s+aprovado|'
    r'busca(?:ndo)?\s+(?:casa|apartamento|apto|imovel|imóvel|terreno)|'
    r'algu[eé]m\s+(?:tem|com|que\s+tenha)\s+\w|'   # "alguém com um X pra venda"
    r'algu[eé]m\s+(?:tem|tem\s+um|sabe\s+de)|'
    r'\bpra\s+venda[,\s].{0,30}(?:precis|quer|busca|procu)',  # "pra venda... preciso"
    re.IGNORECASE)

# "Se vc procura... achou!" = anúncio de venda, não demanda
RE_VENDA = re.compile(
    r'\bvendo\b|\bvende\b|\bà\s+venda\b|\bdisponív|\banuncio\b|\bofereço\b|'
    r'\bchaves\s+na\s+mão\b|\bentrego\s+chaves\b|'
    r'achou[!🎉]|(?:se\s+vc|se\s+você)\s+procura',
    re.IGNORECASE)

def classificar(texto):
    d = bool(RE_DEMANDA.search(texto))
    v = bool(RE_VENDA.search(texto))
    if d and not v: return 'demanda'
    if v and not d: return 'venda'
    if d and v:
        return 'demanda' if RE_DEMANDA.search(texto).start() < RE_VENDA.search(texto).start() else 'venda'
    return 'indefinido'

# ─── Extratores ───────────────────────────────────────────────────────────────

def extrair_preco(texto):
    # Remover caracteres invisíveis (ex: U+2060 WORD JOINER do WhatsApp)
    texto = re.sub(r'[⁠​‌‍﻿]', '', texto)

    # Preço atual quando houve redução: "de R$X para R$Y" → usa Y
    m_red = (
        re.search(r'reduz(?:iu|indo|ão).{0,60}?para\s*R?\$?\s*([\d.,]+)', texto, re.IGNORECASE) or
        re.search(r'de\s+R\$\s*[\d.,]+\s+para\s*R?\$?\s*([\d.,]+)', texto, re.IGNORECASE) or
        re.search(r'baixou.{0,40}?para\s*R?\$?\s*([\d.,]+)', texto, re.IGNORECASE) or
        re.search(r'por\s+apenas\s*R?\$?\s*([\d.,]+)', texto, re.IGNORECASE)
    )
    if m_red:
        raw = m_red.group(1).rstrip('.,')
        try:
            if re.match(r'^\d{1,3}(\.\d{3})+(,\d{2})?$', raw):
                num = float(raw.replace('.','').replace(',','.'))
            else:
                num = float(raw.replace(',',''))
            if 30_000 <= num <= 50_000_000:
                return int(num)
        except: pass

    padroes = [
        (r'R\$\s*([\d.,]+)\s*mi(?:lhão|lhões|l\b)?', 'mi'),
        (r'R\$\s*([\d.,]+)\s*mil\b', 'mil'),
        (r'R\$\s*([\d.,]+)', 'reais'),
        (r'\b(\d+(?:[.,]\d+)?)\s*mi(?:lhão|lhões)\b', 'mi'),   # "1 milhão" sem R$
        (r'\binvestimento[:\s]+(\d+(?:[.,]\d+)?)\s*mil\b', 'mil'),
        (r'\b([\d.,]+)\s*mil\b', 'mil'),
    ]
    for pat, tipo in padroes:
        m = re.search(pat, texto, re.IGNORECASE)
        if not m: continue
        raw = m.group(1).rstrip('.,')  # remove ponto/vírgula final (ex: "2.750.000,00.")
        if not raw: continue
        try:
            if re.match(r'^\d{1,3}(\.\d{3})+(,\d{2})?$', raw):
                num = float(raw.replace('.','').replace(',','.'))
            elif ',' in raw and '.' not in raw:
                num = float(raw.replace(',','.'))
            else:
                num = float(raw.replace(',',''))
            if tipo == 'mi':  num *= 1_000_000
            if tipo == 'mil' and num < 10_000: num *= 1_000
            if 30_000 <= num <= 50_000_000:
                return int(num)
        except: pass
    return None

def extrair_area(texto):
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*m[²2]', texto, re.IGNORECASE)
    if m:
        try: return float(m.group(1).replace(',','.'))
        except: pass
    return None

def extrair_num(texto, palavras):
    for p in palavras:
        m = re.search(r'(\d+)\s*' + p, texto, re.IGNORECASE)
        if m: return int(m.group(1))
    return None

def extrair_tipo(texto):
    t = texto.lower()
    primeira_linha = t.split('\n')[0].strip()

    # Prioridade 1: TÍTULO da primeira linha (ex: "Casa à Venda – ...")
    m_titulo = re.match(
        r'^(casa|sobrado|terreno|lote|apartamento|apto|sala|galpão|kitnet|studio|chácara|sítio)\b',
        primeira_linha)
    if m_titulo:
        p = m_titulo.group(1)
        if 'apart' in p or 'apto' in p:        return 'Apartamento'
        if 'casa'    in p:                      return 'Casa'
        if 'sobrado' in p:                      return 'Sobrado'
        if 'terreno' in p or 'lote' in p:       return 'Terreno'
        if 'sala'    in p:                      return 'Sala Comercial'
        if 'galpão'  in p:                      return 'Galpão'
        if 'kitnet'  in p or 'studio' in p:     return 'Kitnet'
        if 'chácara' in p or 'sítio' in p:      return 'Chácara'

    # Prioridade 2: padrão "proprietária de uma CASA", "vendo uma CASA", etc.
    m_oferta = re.search(
        r'(?:proprietári[ao]\s+de\s+um[a]?\s+|vendo\s+um[a]?\s+|tenho\s+um[a]?\s+|à\s+venda[:\s]+um[a]?\s+)'
        r'(apartamento|apto|casa|terreno|lote|sala|galpão|sobrado|kitnet)',
        t)
    if m_oferta:
        palavra = m_oferta.group(1)
        if 'apart' in palavra or 'apto' in palavra: return 'Apartamento'
        if 'casa'    in palavra: return 'Casa'
        if 'terreno' in palavra or 'lote' in palavra: return 'Terreno'
        if 'sala'    in palavra: return 'Sala Comercial'
        if 'galpão'  in palavra: return 'Galpão'
        if 'sobrado' in palavra: return 'Sobrado'
        if 'kitnet'  in palavra: return 'Kitnet'

    # Sinais de imóvel habitado (quartos, suíte, sala, cozinha) — se presentes,
    # "lote" e "terreno" são apenas medidas, não o tipo do imóvel
    tem_habitacao = bool(re.search(
        r'\bquartos?\b|\bdorm\b|\bsuítes?\b|\bsala\s+(?:de\s+)?(?:estar|jantar|pé\s+direito)\b'
        r'|\bcozinha\b|\bárea\s+privativa\b|\bárea\s+construída\b', t))

    # Prioridade 3: primeira menção no texto completo
    if re.search(r'\bkitnet\b|\bkit\s*net\b|\bstudio\b|\bflat\b', t): return 'Kitnet'
    if re.search(r'\bapartamento|\bapto\b|\bcobertura\b|\bárea\s+privativa\b', t): return 'Apartamento'
    if re.search(r'\bsobrado\b', t):                           return 'Sobrado'
    if re.search(r'\bcasa\b', t):                              return 'Casa'
    # "terreno" e "lote" só classificam como Terreno se não houver sinais de habitação
    if not tem_habitacao and re.search(r'\bterreno\b|\blote\b', t): return 'Terreno'
    if tem_habitacao and re.search(r'\blote\b|\bárea\s+do\s+lote\b', t): return 'Casa'
    if re.search(r'\bsala\s+comercial|\bloja\b|\bescritório\b', t): return 'Sala Comercial'
    if re.search(r'\bgalpão\b', t):                            return 'Galpão'
    if re.search(r'\bedifício|\bed\.\s*[a-zA-Z]|\bandar\b', t): return 'Apartamento'
    return 'Imóvel'

BAIRROS = [
    'Zona 01','Zona 02','Zona 03','Zona 04','Zona 05','Zona 06','Zona 07','Zona 08',
    'Zona 14','Zona 17','Zona 18','Jardim Alvorada','Jardim América','Jardim Astúrias',
    'Jardim Atalaia','Jardim Avenida','Jardim Bela Vista','Jardim Borba Gato',
    'Jardim Catuaí','Jardim Cidade Monções','Jardim Contorno','Jardim Dias',
    'Jardim Dubai','Jardim Europa','Jardim Farolândia','Jardim Finotti',
    'Jardim Florença','Jardim Imperial','Jardim Independência','Jardim Ipanema',
    'Jardim Itaipu','Jardim Liberdade','Jardim Malibu','Jardim Mandacaru',
    'Jardim Mônaco','Jardim Novo Horizonte','Jardim Olímpico','Jardim Panorama',
    'Jardim Paris','Jardim Paulista','Jardim Pinheiros','Jardim Primavera',
    'Jardim Santos Dumont','Jardim São Jorge','Jardim São Paulo','Jardim Sol Nascente',
    'Jardim Tamariz','Jardim Universo','Jardim Vera Cruz','Jardim Vitória','Jardim Yara',
    'Alto da Glória','Alto Alegre','Aeroporto','Centro','Centro Cívico','Floriano',
    'Gleba Palhano','Liberdade','Nova Esperança','Novo Aeroporto',
    'Parque das Laranjeiras','Parque Hortência','Parque Ideal','Santa Felicidade',
    'Santa Cruz','Santa Mônica','Santa Rosa','Santa Terezinha','Tuiuti','Ulyssea',
    'Vigilato Pereira','Palhano','Yara','Morumbi','Jardim Fregadoli',
    'Jardim Gastão Vidigal','Chácaras Aeroporto','Vila Operária','Vila Morangueira',
    # Loteamentos e bairros menos conhecidos
    'Jardim 3 Lagoas','Jardim Três Lagoas','Jardim Universitário','Jardim Altos do Mirante',
    'Jardim Colinas','Jardim Copacabana','Jardim Dinalva','Jardim Dom Bosco',
    'Jardim Flamingos','Jardim Francos','Jardim Ibiporã','Jardim Marcelo',
    'Jardim Monte Rei','Jardim Nobre','Jardim Olimpo','Jardim Ouro Branco',
    'Jardim Suíço','Jardim Tuiuti','Parque Alvorada','Parque Estação',
    'Residencial Cidade Universitária','Cidade Universitária',
    'Conjunto Residencial Requião','Jardim Requião',
]
BAIRROS_LOWER = {b.lower(): b for b in BAIRROS}

# ─── Cache de locais (bairro vs condomínio) ──────────────────────────────────
CACHE_LOCAIS_FILE = BASE_DIR / "cache_locais.json"

def _carregar_cache_locais():
    if CACHE_LOCAIS_FILE.exists():
        try: return json.load(open(CACHE_LOCAIS_FILE, encoding='utf-8'))
        except: pass
    return {}

def _salvar_cache_locais(cache):
    with open(CACHE_LOCAIS_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

def classificar_local(nome):
    """
    Usa Claude Haiku para descobrir se um nome é bairro ou condomínio/edifício.
    Resultado armazenado em cache_locais.json para não repetir a consulta.
    Retorna dict: {'tipo': 'bairro'|'condominio'|'outro', 'nome': str, 'bairro_real': str|None}
    Efeito colateral: se for condomínio novo, adiciona a _CONDOS_NOVOS para pesquisa posterior.
    """
    if not nome or len(nome.strip()) < 3:
        return {'tipo': 'outro', 'nome': nome, 'bairro_real': None}

    cache = _carregar_cache_locais()
    chave = nome.strip().lower()
    if chave in cache:
        resultado = cache[chave]
        # Mesmo em cache: se é condomínio, verificar se ainda não foi pesquisado
        if resultado.get('tipo') == 'condominio' and f"_pesq_{chave}" not in cache:
            _CONDOS_NOVOS.add(resultado.get('nome', nome))
        return resultado

    api_key = _api_key()
    if not api_key:
        return {'tipo': 'outro', 'nome': nome, 'bairro_real': None}

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            f'Em Maringá-PR, o nome "{nome}" se refere a:\n'
            f'1) Um bairro ou zona oficial da cidade\n'
            f'2) Um condomínio, edifício ou empreendimento imobiliário\n'
            f'3) Outro (cidade, rua, etc)\n\n'
            f'Se for condomínio/edifício, em qual bairro de Maringá fica?\n\n'
            f'Responda SOMENTE JSON válido:\n'
            f'{{"tipo":"bairro"|"condominio"|"outro",'
            f'"nome":"nome mais completo/oficial se souber",'
            f'"bairro_real":"bairro onde fica (se condomínio) ou null"}}'
        )
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=150,
            messages=[{"role": "user", "content": prompt}]
        )
        m = re.search(r'\{.*\}', resp.content[0].text, re.DOTALL)
        if m:
            resultado = json.loads(m.group())
            cache[chave] = resultado
            _salvar_cache_locais(cache)
            tipo = resultado.get('tipo', 'outro')
            nome_resultado = resultado.get('nome', nome)
            print(f"  🗺️  '{nome}' → {tipo.upper()}: {nome_resultado}"
                  + (f" (bairro: {resultado['bairro_real']})" if resultado.get('bairro_real') else ""))
            # Se for condomínio novo → agendar pesquisa detalhada
            if tipo == 'condominio':
                _CONDOS_NOVOS.add(nome_resultado)
            return resultado
    except Exception as e:
        print(f"  ⚠️  classificar_local: {e}")

    resultado = {'tipo': 'outro', 'nome': nome, 'bairro_real': None}
    cache[chave] = resultado
    _salvar_cache_locais(cache)
    return resultado


# ─── Pesquisa de condomínios (web search via Claude) ─────────────────────────

def _condos_ja_na_planilha():
    """Retorna set com nomes (lowercase) dos condomínios já cadastrados."""
    if not PLANILHA.exists():
        return set()
    try:
        df = pd.read_excel(PLANILHA, sheet_name=ABA_CONDOMINIOS)
        return {str(n).lower().strip() for n in df['Nome'].dropna()}
    except:
        return set()


def pesquisar_condominio(nome, cidade="Maringá-PR"):
    """
    Pesquisa dados completos de um condomínio via Claude Sonnet + web_search.
    Retorna dict com informações ou None se já pesquisado/erro.
    """
    api_key = _api_key()
    if not api_key:
        return None

    cache = _carregar_cache_locais()
    chave_pesq = f"_pesq_{nome.strip().lower()}"
    if chave_pesq in cache:
        print(f"  🏗️  '{nome}' já pesquisado anteriormente (cache)")
        return None

    print(f"  🔎 Pesquisando condomínio '{nome}' em {cidade}...")

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = (
            f'Pesquise o empreendimento/condomínio "{nome}" em {cidade}.\n'
            f'Quero informações completas para cadastro imobiliário.\n\n'
            f'Retorne SOMENTE JSON válido (sem markdown, sem texto extra):\n'
            f'{{\n'
            f'  "nome": "nome completo oficial",\n'
            f'  "endereco": "rua e número",\n'
            f'  "bairro": "bairro em Maringá",\n'
            f'  "cep": "00000-000 ou null",\n'
            f'  "construtora": "nome da construtora/incorporadora",\n'
            f'  "ano_lancamento": "YYYY ou null",\n'
            f'  "previsao_entrega": "YYYY ou null",\n'
            f'  "padrao": "Econômico|Médio Padrão|Alto Padrão|Luxo",\n'
            f'  "torres": "número de torres ou null",\n'
            f'  "andares": "número de andares ou null",\n'
            f'  "total_aptos": "total de apartamentos ou null",\n'
            f'  "area_min": número_em_m2_ou_null,\n'
            f'  "area_max": número_em_m2_ou_null,\n'
            f'  "quartos": "ex: 2 e 3 quartos",\n'
            f'  "vagas": "ex: 1 a 2 vagas",\n'
            f'  "lazer": "lista separada por vírgula: piscina, academia, salão...",\n'
            f'  "faixa_preco": "ex: R$350.000 a R$550.000",\n'
            f'  "observacoes": "informações adicionais relevantes",\n'
            f'  "link": "URL principal do empreendimento"\n'
            f'}}'
        )

        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1500,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
            messages=[{"role": "user", "content": prompt}]
        )

        # Extrair texto (ignorar tool_use blocks)
        text = ""
        for block in resp.content:
            if hasattr(block, 'text'):
                text += block.text

        m = re.search(r'\{[\s\S]*\}', text)
        if m:
            info = json.loads(m.group())
            # Marcar como pesquisado no cache
            cache[chave_pesq] = True
            _salvar_cache_locais(cache)
            print(f"  ✅ Dados obtidos para '{info.get('nome', nome)}'")
            return info

    except Exception as e:
        print(f"  ⚠️  pesquisar_condominio('{nome}'): {e}")

    return None


def atualizar_aba_condominios(info):
    """Insere condomínio na aba Condomínios se ainda não estiver lá."""
    from datetime import datetime

    nome = str(info.get('nome', '') or '').strip()
    if not nome:
        return

    # Verificar se já está na planilha
    ja_cadastrados = _condos_ja_na_planilha()
    if nome.lower() in ja_cadastrados:
        print(f"  ⏭️  '{nome}' já está na aba {ABA_CONDOMINIOS}")
        return

    linha = [
        nome,
        info.get('endereco') or '',
        info.get('bairro') or '',
        info.get('cep') or '',
        info.get('construtora') or '',
        info.get('ano_lancamento') or '',
        info.get('previsao_entrega') or '',
        info.get('padrao') or '',
        info.get('torres') or '',
        info.get('andares') or '',
        info.get('total_aptos') or '',
        info.get('area_min'),
        info.get('area_max'),
        info.get('quartos') or '',
        info.get('vagas') or '',
        info.get('lazer') or '',
        info.get('faixa_preco') or '',
        info.get('observacoes') or '',
        info.get('link') or '',
        datetime.now().strftime('%d/%m/%Y %H:%M'),
    ]

    inserir_linhas(ABA_CONDOMINIOS, [linha], COLUNAS_CONDOMINIOS)
    print(f"  🏗️  Condomínio '{nome}' cadastrado na aba {ABA_CONDOMINIOS}")

# Nomes geográficos que NÃO são bairros (cidade, estado, país)
_NAO_BAIRRO = {
    'maringá', 'maringa', 'londrina', 'curitiba', 'são paulo', 'sao paulo',
    'brasil', 'brazil', 'paraná', 'parana', 'pr',
}

def _expandir_abreviaturas(texto):
    """Expande abreviações comuns de bairros para facilitar a busca."""
    t = texto
    t = re.sub(r'\bJD\.\s*', 'Jardim ', t, flags=re.IGNORECASE)
    t = re.sub(r'\bRES\.\s*', 'Residencial ', t, flags=re.IGNORECASE)
    t = re.sub(r'\bCOND\.\s*', 'Condomínio ', t, flags=re.IGNORECASE)
    t = re.sub(r'\bPQ\.\s*', 'Parque ', t, flags=re.IGNORECASE)
    t = re.sub(r'\bAV\.\s*', 'Avenida ', t, flags=re.IGNORECASE)
    return t

def extrair_bairro(texto):
    # Expandir abreviações antes de comparar
    texto_exp = _expandir_abreviaturas(texto)
    tl = texto_exp.lower()

    # 0. Padrão do título: "Tipo à Venda – Bairro | Cidade" (ex: "Casa à Venda – Jardim Fregadoli | Maringá")
    m_titulo = re.search(
        r'(?:venda|aluguel|locação)\s*[–\-—]+\s*([A-ZÀ-Ú][A-Za-zÀ-ú\s]{2,40}?)\s*\|',
        texto_exp, re.IGNORECASE)
    if m_titulo:
        candidato = m_titulo.group(1).strip()
        if candidato.lower() not in _NAO_BAIRRO and 2 < len(candidato) < 45:
            # Verificar se está na lista de bairros
            if candidato.lower() in BAIRROS_LOWER:
                return BAIRROS_LOWER[candidato.lower()]
            # Aceitar diretamente como bairro (está no título, alta confiança)
            return candidato

    # 1. Verificar lista de bairros conhecidos (com texto expandido)
    for bl, b in BAIRROS_LOWER.items():
        if bl in tl: return b
    # 2. Tentar extrair nome por padrão contextual (texto já expandido)
    m = re.search(r'(?:no|na|em|bairro|condomínio|cond\.?|edifício|ed\.?|residencial|região)\s+([A-ZÀ-Ú0-9][a-zA-ZÀ-ú0-9\s]{2,35}?)(?:\s*[-–,.]|\s*$|\s*\n)', texto_exp)
    if m:
        c = m.group(1).strip()
        # Filtrar nomes de cidades/estados que não são bairros
        if c.lower() in _NAO_BAIRRO:
            return ''
        if 2 < len(c) < 40:
            # Verificar se esse nome está na lista de bairros
            if c.lower() in BAIRROS_LOWER:
                return BAIRROS_LOWER[c.lower()]
            # Se não, classificar via IA (resultado fica em cache)
            info = classificar_local(c)
            if info['tipo'] == 'bairro':
                return info.get('nome', c)
            elif info['tipo'] == 'condominio':
                bairro_real = info.get('bairro_real')
                if bairro_real:
                    return f"Cond. {info.get('nome', c)} · {bairro_real}"
                return f"Cond. {info.get('nome', c)}"
    return ''

def extrair_edificio(texto):
    """Extrai nome de edifício/condomínio mencionado explicitamente no texto."""
    m = re.search(
        r'(?:edifício|ed\.|condomínio|cond\.|residencial)\s+([A-ZÀ-Ú][A-Za-zÀ-ú\s]{2,35}?)(?:\s*[·\-,\.]|\s*\d+[oOºª]|\s*$|\n)',
        texto, re.IGNORECASE
    )
    if m:
        nome = m.group(1).strip().rstrip('·-.,')
        if 2 < len(nome) < 40:
            return nome
    return None

def extrair_campos(texto):
    edificio = extrair_edificio(texto)
    # Se achou nome de edifício → acionar classificação/pesquisa
    if edificio:
        info = classificar_local(edificio)
        # se for condomínio novo, já foi adicionado a _CONDOS_NOVOS dentro de classificar_local
    return {
        'tipo':      extrair_tipo(texto),
        'bairro':    extrair_bairro(texto),
        'area':      extrair_area(texto),
        'quartos':   extrair_num(texto, [r'quartos?', r'dormit[oó]rios?', r'dorm\.?']),
        'suites':    extrair_num(texto, [r'su[íi]tes?']),
        'banheiros': extrair_num(texto, [r'banheiros?', r'\bwc\b', r'lavabo']),
        'vagas':     extrair_num(texto, [r'vagas?', r'garagens?']),
        'preco':     extrair_preco(texto),
    }

def tem_dados(c):
    return any([c.get('preco'), c.get('area'), c.get('quartos'), c.get('suites'), c.get('vagas')])

# ─── Agrupamento: fotos + texto do mesmo corretor = 1 imóvel ─────────────────

def agrupar_mensagens(pendentes):
    """
    Agrupa mensagens próximas no tempo do mesmo autor no mesmo grupo.
    Retorna lista de pacotes — cada pacote = 1 imóvel.
    """
    if not pendentes:
        return []

    ordenadas = sorted(enumerate(pendentes), key=lambda x: x[1].get('timestamp', 0))
    usadas = set()
    pacotes = []

    for idx_orig, msg in ordenadas:
        if idx_orig in usadas:
            continue

        ts = msg.get('timestamp', 0)
        pacote_idxs = [idx_orig]
        usadas.add(idx_orig)

        for idx2, outro in ordenadas:
            if idx2 in usadas:
                continue
            if (outro['autor'] == msg['autor'] and
                outro['grupo'] == msg['grupo'] and
                abs(outro.get('timestamp', 0) - ts) <= JANELA_AGRUPAMENTO):
                pacote_idxs.append(idx2)
                usadas.add(idx2)

        pacotes.append({
            'idxs': pacote_idxs,
            'msgs': [pendentes[i] for i in pacote_idxs],
            'autor': msg['autor'],
            'grupo': msg['grupo'],
            'data':  msg.get('data', ''),
            'contato': msg.get('contato', ''),
        })

    return pacotes

def resolver_pacote(pacote):
    """
    De um pacote de mensagens (texto + fotos), extrai os dados do imóvel.
    Regra: se o texto tem dados → usa o texto, ignora imagens.
           se só tem imagens → analisa UMA imagem com Claude.
    Retorna (campos, obs, classe) ou None se não for imóvel.
    """
    msgs = pacote['msgs']

    # Juntar todo o texto do pacote
    textos = [m.get('texto', '') for m in msgs if m.get('texto')]
    texto_completo = '\n'.join(textos).strip()

    # Classificar pelo texto
    classe = classificar(texto_completo) if texto_completo else 'indefinido'

    # Extrair campos do texto
    campos = extrair_campos(texto_completo) if texto_completo else None

    if campos and tem_dados(campos):
        # Texto tem dados suficientes → não precisa analisar imagem
        obs = texto_completo[:300]
        return campos, obs, classe

    # Texto insuficiente → tentar UMA imagem (a primeira com arquivo salvo)
    img_msgs = [m for m in msgs if m.get('imagemPath') and Path(m['imagemPath']).exists()]
    if img_msgs:
        img_msg = img_msgs[0]  # só analisar a primeira
        n_imgs = len(img_msgs)
        print(f"  🔍 Claude analisa 1 imagem de {n_imgs} [{pacote['autor']}]")
        resultado = analisar_imagem(img_msg['imagemPath'], texto_completo, pacote['autor'])
        if resultado and resultado.get('eh_imovel'):
            campos = {
                'tipo':      resultado.get('tipo', 'Imóvel'),
                'bairro':    resultado.get('bairro') or '',
                'area':      resultado.get('area'),
                'quartos':   resultado.get('quartos'),
                'suites':    resultado.get('suites'),
                'banheiros': resultado.get('banheiros'),
                'vagas':     resultado.get('vagas'),
                'preco':     resultado.get('preco'),
            }
            # Só inserir se tiver pelo menos 1 dado concreto (preço, área, quartos...)
            if not tem_dados(campos):
                print(f"     ⏭️  Imagem de imóvel sem dados concretos — ignorando")
                return None
            obs = resultado.get('obs', '') or texto_completo[:300]
            if classe == 'indefinido':
                classe = 'venda'
            print(f"     ✅ {campos['tipo']} | {campos.get('bairro') or '?'} | R${campos.get('preco')}")
            return campos, obs, classe
        # Claude confirmou que não é imóvel, ou não conseguiu analisar → pular
        return None

    # Imagem sem arquivo local (download falhou ou não suportado) → pular sem criar placeholder
    return None  # sem dados suficientes

# ─── Deduplicação ─────────────────────────────────────────────────────────────

def fazer_fp(autor, bairro, preco, area, texto, timestamp=None):
    autor = str(autor or '').lower().strip()
    bairro = str(bairro or '').lower().strip()
    preco = preco or 0
    area = area or 0
    texto_curto = str(texto or '')[:80].lower().strip()

    if bairro and preco:       return f"{autor}|{bairro}|{preco}"
    elif preco and area:       return f"{autor}|{preco}|{area}"
    elif preco:                return f"{autor}|{preco}"
    elif texto_curto:          return f"{autor}|txt:{texto_curto}"
    elif timestamp:            return f"{autor}|ts:{int(timestamp)}"  # fallback: autor+timestamp
    return None  # último recurso: não deduplica

def _load_fps(sheet_name, col_autor, col_bairro, col_preco, col_area, col_obs):
    if not PLANILHA.exists(): return set()
    try:
        df = pd.read_excel(PLANILHA, sheet_name=sheet_name)
        fps = set()
        for _, r in df.iterrows():
            fp = fazer_fp(
                r.get(col_autor,''), r.get(col_bairro,''),
                r.get(col_preco), r.get(col_area), r.get(col_obs,''))
            if fp: fps.add(fp)
        return fps
    except: return set()

def fp_imoveis():
    return _load_fps(ABA_IMOVEIS, 'Corretor', 'Bairro / Endereço', 'Preço (R$)', 'Área (m²)', 'Observações')

def fp_demandas():
    return _load_fps(ABA_DEMANDAS, 'Corretor', 'Bairro/Região', 'Orçamento Máx', 'Área Mín', 'Observações')

# ─── Planilha ─────────────────────────────────────────────────────────────────

def inserir_linhas(aba, linhas, colunas):
    if not PLANILHA.exists():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active; ws.title = aba; ws.append(colunas)
        wb.save(PLANILHA)

    wb = load_workbook(PLANILHA)
    if aba not in wb.sheetnames:
        ws = wb.create_sheet(aba); ws.append(colunas)
    else:
        ws = wb[aba]

    cor = PatternFill(fill_type='solid', fgColor='FFF9C4')
    for linha in linhas:
        ws.append(linha)
        for cell in ws[ws.max_row]: cell.fill = cor
    wb.save(PLANILHA)

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not FILA_FILE.exists():
        print("Fila vazia — bot Baileys ainda não capturou mensagens.")
        return

    with open(FILA_FILE, 'r', encoding='utf-8') as f:
        fila = json.load(f)

    pendentes = [m for m in fila if not m.get('processado')]
    print(f"📬 {len(pendentes)} mensagens pendentes (total: {len(fila)})\n")

    if VER_FILA:
        for i, m in enumerate(pendentes, 1):
            cls = classificar(m.get('texto',''))
            ic = '🏠' if cls=='venda' else '🔍' if cls=='demanda' else '❓'
            print(f"── {i}. {ic} [{m['grupo']}] {m['autor']} ({m.get('data','')})")
            img = f" + 🖼️" if m.get('temImagem') else ""
            print(f"   {m.get('texto','(imagem)')[:150]}{img}\n")
        return

    # Agrupar por autor + grupo + tempo
    pacotes = agrupar_mensagens(pendentes)
    print(f"📦 {len(pacotes)} pacotes agrupados (era {len(pendentes)} msgs individuais)\n")

    fps_v = fp_imoveis()
    fps_d = fp_demandas()
    novas_vendas, novas_demandas = [], []
    sem_dados = duplicatas = 0

    for pacote in pacotes:
        resultado = resolver_pacote(pacote)

        # Marcar todas as mensagens do pacote como processadas
        for idx in pacote['idxs']:
            fila[idx]['processado'] = True

        if resultado is None:
            sem_dados += 1
            continue

        campos, obs, classe = resultado

        ts0 = pacote['msgs'][0].get('timestamp') if pacote['msgs'] else None

        if classe == 'demanda':
            fp = fazer_fp(pacote['autor'], campos.get('bairro',''), campos.get('preco'), campos.get('area'), obs, ts0)
            if fp and fp in fps_d:
                duplicatas += 1
                continue
            # Validar contato: LIDs (>13 dígitos) não são telefones reais
            contato_raw = str(pacote.get('contato') or '').replace('.', '').replace(' ', '')
            contato_ok = contato_raw if (contato_raw.isdigit() and 10 <= len(contato_raw) <= 13) else ''
            linha = [
                pacote['data'], pacote['grupo'], pacote['autor'], contato_ok,
                campos['tipo'], campos.get('bairro',''), campos.get('area'),
                campos.get('quartos'), campos.get('suites'), campos.get('banheiros'),
                campos.get('vagas'), campos.get('preco'), obs, 'Nova'
            ]
            novas_demandas.append(linha)
            fps_d.add(fp or f"{pacote['autor']}|ts:{ts0}")

            if DRY_RUN:
                n = len(pacote['msgs'])
                print(f"🔍 DEMANDA ({n} msgs) | {campos['tipo']} | {campos.get('bairro') or '?'} | orç. R${campos.get('preco')}")
                print(f"   [{pacote['grupo']}] {pacote['autor']}\n")

        else:  # venda ou indefinido
            fp = fazer_fp(pacote['autor'], campos.get('bairro',''), campos.get('preco'), campos.get('area'), obs, ts0)
            if fp and fp in fps_v:
                duplicatas += 1
                continue
            contato_raw = str(pacote.get('contato') or '').replace('.', '').replace(' ', '')
            contato_ok = contato_raw if (contato_raw.isdigit() and 10 <= len(contato_raw) <= 13) else ''
            linha = [
                pacote['data'], pacote['grupo'], pacote['autor'], contato_ok,
                campos['tipo'], campos.get('bairro',''), campos.get('area'),
                campos.get('quartos'), campos.get('suites'), campos.get('banheiros'),
                campos.get('vagas'), campos.get('preco'), obs, 'Novo', ''
            ]
            novas_vendas.append(linha)
            fps_v.add(fp or f"{pacote['autor']}|ts:{ts0}")

            if DRY_RUN:
                n = len(pacote['msgs'])
                print(f"🏠 VENDA ({n} msgs→1) | {campos['tipo']} | {campos.get('bairro') or '?'} | "
                      f"{campos.get('area')}m² | R${campos.get('preco')}")
                print(f"   [{pacote['grupo']}] {pacote['autor']}\n")

    if not DRY_RUN:
        if novas_vendas:
            inserir_linhas(ABA_IMOVEIS, novas_vendas, COLUNAS_IMOVEIS)
        if novas_demandas:
            inserir_linhas(ABA_DEMANDAS, novas_demandas, COLUNAS_DEMANDAS)

        with open(FILA_FILE, 'w', encoding='utf-8') as f:
            json.dump(fila, f, ensure_ascii=False, indent=2)

    print(f"\n{'[DRY-RUN] ' if DRY_RUN else ''}✅ {len(novas_vendas)} imóveis inseridos → '{ABA_IMOVEIS}'")
    print(f"{'[DRY-RUN] ' if DRY_RUN else ''}🔍 {len(novas_demandas)} demandas inseridas → '{ABA_DEMANDAS}'")
    print(f"   ↳ {duplicatas} duplicatas ignoradas")
    print(f"   ↳ {sem_dados} pacotes sem dados de imóvel")

    # ── Pesquisar e cadastrar condomínios novos encontrados nesta execução ──────
    if not DRY_RUN and _CONDOS_NOVOS:
        ja_na_planilha = _condos_ja_na_planilha()
        novos_para_pesquisar = [n for n in _CONDOS_NOVOS if n.lower() not in ja_na_planilha]
        if novos_para_pesquisar:
            print(f"\n🏗️  Pesquisando {len(novos_para_pesquisar)} condomínio(s) novo(s)...")
            for nome_condo in novos_para_pesquisar:
                info = pesquisar_condominio(nome_condo)
                if info:
                    atualizar_aba_condominios(info)

if __name__ == '__main__':
    main()
