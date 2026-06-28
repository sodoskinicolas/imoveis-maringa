#!/usr/bin/env python3
"""
scrape_vivareal.py
Extrai imóveis à venda em Maringá do VivaReal.

Uso:
    pip install requests openpyxl
    python scrape_vivareal.py                    # 10 páginas (~240 imóveis)
    python scrape_vivareal.py --paginas 30       # mais imóveis
"""

import re
import sys
import time
import argparse
from pathlib import Path
from datetime import datetime

try:
    import requests
except ImportError:
    print("Instale com: pip install requests openpyxl")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale com: pip install openpyxl")
    sys.exit(1)

# ── Configuração ─────────────────────────────────────────────────────────────
PLANILHA   = Path(__file__).parent / "Imoveis_Grupos.xlsx"
VIVAREAL_XLSX = Path(__file__).parent / "VivaReal_Imoveis.xlsx"
BASE_SEARCH = "https://www.vivareal.com.br/venda/parana/maringa/?pagina={}"
BASE_URL    = "https://www.vivareal.com.br"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
}

TIPO_MAP = {
    "apartamento": "Apartamento", "casa": "Casa", "lote": "Terreno",
    "terreno": "Terreno", "sobrado": "Sobrado", "cobertura": "Cobertura",
    "sala": "Sala Comercial", "ponto": "Ponto Comercial", "galpao": "Galpão",
    "kitnet": "Kitnet", "studio": "Studio", "chacara": "Chácara/Sítio",
    "fazenda": "Chácara/Sítio", "imovel": "Imóvel Comercial",
}

MESES = {
    "janeiro":"01","fevereiro":"02","março":"03","abril":"04",
    "maio":"05","junho":"06","julho":"07","agosto":"08",
    "setembro":"09","outubro":"10","novembro":"11","dezembro":"12",
}

URL_RE  = re.compile(r'href="(https://www\.vivareal\.com\.br/imovel/[^"]+id-(\d+)[^"]*)"')
DATE_RE = re.compile(r'Anúncio criado em (\d+) de (\w+) de (\d{4})')

# ── Helpers ──────────────────────────────────────────────────────────────────
def _int(text, pat):
    m = re.search(pat, text)
    return int(m.group(1)) if m else None

def _float(text, pat):
    m = re.search(pat, text)
    if m:
        try: return float(m.group(1).replace(".","").replace(",","."))
        except: return None
    return None

def parse_url_slug(url):
    slug = url.split("/imovel/")[-1].split("?")[0].rstrip("/")
    parts = slug.split("-")
    tipo  = TIPO_MAP.get(parts[0], parts[0].capitalize())
    area  = next((int(p[:-2]) for p in parts if p.endswith("m2") and p[:-2].isdigit()), None)
    preco = next((int(p[2:]) for p in parts if p.startswith("RS") and p[2:].isdigit()), None)
    quartos = None
    for i, p in enumerate(parts):
        if p == "quartos" and i > 0 and parts[i-1].isdigit():
            quartos = int(parts[i-1]); break
    return {"tipo": tipo, "area": area, "preco": preco, "quartos": quartos}

def parse_search_page(html):
    """Extrai URLs únicas de imóveis de uma página de busca."""
    found = {}
    for url, id_ in URL_RE.findall(html):
        clean = url.split("?")[0]
        if id_ not in found:
            found[id_] = clean
    return list(found.items())   # [(id, url), ...]

def fetch_listing(session, id_, url):
    """Busca página individual e extrai todos os dados incluindo data."""
    try:
        time.sleep(0.7)
        r = session.get(url, timeout=12)
        t = r.text

        # Data de publicação
        dm = DATE_RE.search(t)
        data_pub = ""
        if dm:
            d, m, a = dm.groups()
            data_pub = f"{a}-{MESES.get(m.lower(),'01')}-{d.zfill(2)}"

        # Endereço: "Rua X, 123 - Bairro, Maringá - PR"
        addr = re.search(r'([^,\n]+),\s*([^-\n]+)\s*-\s*Maringá\s*-\s*PR', t)
        rua, bairro = ("", "")
        if addr:
            rua    = addr.group(1).strip()
            bairro = addr.group(2).strip()

        # Campos numéricos
        area    = _float(t, r'(\d[\d.,]*)\s*m²')
        quartos = _int(t,  r'(\d+)\s*quarto')
        suites  = _int(t,  r'(\d+)\s*suíte')
        banhs   = _int(t,  r'(\d+)\s*banheiro')
        vagas   = _int(t,  r'(\d+)\s*vaga')
        preco   = _float(t, r'R\$\s*([\d.,]+)')

        # Corretor / imobiliária
        cor_m = re.search(r'(?:Contatar|Chamar|com)\s+([A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ][A-Za-záéíóúâêîôûãõçÁÉÍÓÚÂÊÎÔÛÃÕÇ ]{3,60}?)(?:\s*\n|\s*WhatsApp|\s*Imóveis)', t)
        corretor = cor_m.group(1).strip() if cor_m else ""

        return {
            "id": id_, "link": url,
            "data_pub": data_pub,
            "bairro": bairro, "rua": rua,
            "area": area, "quartos": quartos, "suites": suites,
            "banheiros": banhs, "vagas": vagas,
            "preco": int(preco) if preco else None,
            "corretor": corretor,
        }
    except Exception as e:
        print(f"    ⚠ Erro em {url}: {e}")
        return {"id": id_, "link": url}

# ── Scraper principal ─────────────────────────────────────────────────────────
def scrape(max_pages=10):
    session = requests.Session()
    session.headers.update(HEADERS)
    all_ids = {}   # id → url (dedup)

    # 1. Coleta URLs de todas as páginas de busca
    print(f"🔍 Coletando URLs ({max_pages} páginas)…")
    for page in range(1, max_pages + 1):
        url = BASE_SEARCH.format(page)
        print(f"  Página {page}/{max_pages}  {url}")
        try:
            r = session.get(url, timeout=15)
            pairs = parse_search_page(r.text)
            for id_, u in pairs:
                all_ids.setdefault(id_, u)
            print(f"    → {len(pairs)} imóveis | total acumulado: {len(all_ids)}")
        except Exception as e:
            print(f"  ⚠ Erro: {e}")
        time.sleep(1.2)

    # 2. Visita cada imóvel para extrair detalhes e data
    listings = []
    total = len(all_ids)
    print(f"\n📋 Extraindo detalhes de {total} imóveis…")
    for i, (id_, url) in enumerate(all_ids.items(), 1):
        print(f"  [{i:3}/{total}] {url.split('/')[-2][:55]}")
        basic   = parse_url_slug(url)
        details = fetch_listing(session, id_, url)
        listing = {
            "id":               id_,
            "data_publicacao":  details.get("data_pub", ""),
            "data_captura":     datetime.now().strftime("%Y-%m-%d"),
            "tipo":             details.get("tipo") or basic.get("tipo", ""),
            "bairro":           details.get("bairro", ""),
            "rua":              details.get("rua", ""),
            "area":             details.get("area") or basic.get("area"),
            "quartos":          details.get("quartos") or basic.get("quartos"),
            "suites":           details.get("suites"),
            "banheiros":        details.get("banheiros"),
            "vagas":            details.get("vagas"),
            "preco":            details.get("preco") or basic.get("preco"),
            "corretor":         details.get("corretor", ""),
            "link":             url,
        }
        listings.append(listing)

    return listings

# ── Salvar na planilha padrão (Imoveis_Grupos.xlsx) ──────────────────────────
def salvar_planilha_padrao(listings):
    if not PLANILHA.exists():
        print(f"⚠ Planilha padrão não encontrada: {PLANILHA}")
        return

    wb = load_workbook(PLANILHA)
    ws = wb["Imóveis"]

    # Garantir que a coluna "Data Publicação" existe
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 2)]
    if "Data Publicação" not in headers:
        col_idx = ws.max_column + 1
        ws.cell(1, col_idx).value = "Data Publicação"
        ws.cell(1, col_idx).font  = Font(bold=True, color="FFFFFF")
        ws.cell(1, col_idx).fill  = PatternFill("solid", start_color="1F4E79")
    else:
        col_idx = headers.index("Data Publicação") + 1

    hoje = datetime.now().strftime("%Y-%m-%d")
    inseridos = 0
    # Checa duplicatas por link
    existing_links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        obs = str(row[11] or "")    # coluna Observações
        existing_links.add(obs)

    for im in listings:
        # Verifica se já existe pelo ID no link
        link = im["link"]
        id_  = im["id"]
        if any(id_ in lk for lk in existing_links):
            continue
        obs_val = f"{im['tipo']} {im['bairro']} | id:{id_}"
        ws.append([
            hoje,                           # Data Captura
            "vivareal.com.br",              # Grupo
            im["corretor"] or "VivaReal",   # Corretor
            "",                             # Contato (WhatsApp) - não disponível
            im["tipo"],                     # Tipo
            f"{im['bairro']} · {im['rua']}" if im["rua"] else im["bairro"],  # Bairro/End.
            im["area"],                     # Área
            im["quartos"],                  # Quartos
            im["suites"],                   # Suítes
            im["vagas"],                    # Vagas
            im["preco"],                    # Preço
            obs_val,                        # Observações
            "Venda",                        # Status
            im.get("data_publicacao",""),   # Data Publicação (nova coluna)
        ])
        existing_links.add(obs_val)
        inseridos += 1

    wb.save(PLANILHA)
    print(f"✅ {inseridos} imóveis adicionados em {PLANILHA}")

# ── Salvar planilha separada VivaReal_Imoveis.xlsx ────────────────────────────
def salvar_vivareal_xlsx(listings):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VivaReal Maringá"

    cols = ["ID VivaReal","Data Publicação","Data Captura","Tipo","Bairro","Endereço",
            "Área (m²)","Quartos","Suítes","Banheiros","Vagas","Preço (R$)","Corretor","Link"]

    # Cabeçalho
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(1, ci, c)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18

    for im in listings:
        ws.append([
            im["id"],
            im.get("data_publicacao",""),
            im.get("data_captura",""),
            im.get("tipo",""),
            im.get("bairro",""),
            im.get("rua",""),
            im.get("area"),
            im.get("quartos"),
            im.get("suites"),
            im.get("banheiros"),
            im.get("vagas"),
            im.get("preco"),
            im.get("corretor",""),
            im.get("link",""),
        ])

    # Ajuste de largura
    widths = [16,16,14,16,22,30,10,8,8,10,8,14,28,60]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(VIVAREAL_XLSX)
    print(f"✅ {len(listings)} imóveis salvos em {VIVAREAL_XLSX}")

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scrape VivaReal Maringá")
    parser.add_argument("--paginas", type=int, default=10,
                        help="Número de páginas de busca (24 imóveis/página, default=10)")
    args = parser.parse_args()

    print(f"🏠 VivaReal Scraper — Maringá")
    print(f"   Páginas: {args.paginas} (~{args.paginas*24} imóveis + visitas individuais para datas)")
    print()

    listings = scrape(args.paginas)
    print(f"\n✅ {len(listings)} imóveis extraídos\n")
    salvar_vivareal_xlsx(listings)
    salvar_planilha_padrao(listings)
    print("\nPronto! Rode python gerar_site.py para atualizar o site.")
