#!/usr/bin/env python3
"""
bot_grupos_wa.py
Recebe dados extraídos pelo bot WhatsApp (JSON) e insere na planilha Imoveis_Grupos.xlsx
sem duplicar registros.

Uso:
  python bot_grupos_wa.py --dados '{"grupo":"Maringá APTs","corretor":"João","contato":"554499...",
                                    "tipo":"Apartamento","bairro":"Zona 7","area":"80",
                                    "quartos":"3","suites":"1","vagas":"2","preco":"420000",
                                    "obs":"Sacada gourmet"}'

  python bot_grupos_wa.py --arquivo dados_capturados.json
"""

import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

PLANILHA = Path(__file__).parent / "Imoveis_Grupos.xlsx"

COL_ORDER = [
    "data_captura", "grupo", "corretor", "contato",
    "tipo", "bairro", "area", "quartos",
    "suites", "vagas", "preco", "obs", "status"
]

ALT_FILL   = PatternFill("solid", start_color="DEEAF1")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
BORDER     = Border(
    left=Side(style="thin", color="B0C4D8"),
    right=Side(style="thin", color="B0C4D8"),
    top=Side(style="thin", color="B0C4D8"),
    bottom=Side(style="thin", color="B0C4D8"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP   = Alignment(vertical="center", wrap_text=True)


def normalizar(item: dict) -> dict:
    """Garante campos padrão e formata data."""
    out = {k: "" for k in COL_ORDER}
    out.update({
        "data_captura": datetime.now().strftime("%Y-%m-%d"),
        "status":       "Novo",
    })
    mapa = {
        # aceita variações de chave enviadas pelo bot
        "grupo": ["grupo", "group", "nome_grupo"],
        "corretor": ["corretor", "broker", "nome"],
        "contato": ["contato", "telefone", "phone", "whatsapp", "numero"],
        "tipo": ["tipo", "type", "tipo_imovel"],
        "bairro": ["bairro", "endereco", "address", "localizacao", "location"],
        "area": ["area", "m2", "metragem", "tamanho"],
        "quartos": ["quartos", "bedrooms", "dorms", "dormitorios"],
        "suites": ["suites", "suite", "suítes"],
        "vagas": ["vagas", "garagem", "garage"],
        "preco": ["preco", "preco_venda", "valor", "price"],
        "obs": ["obs", "observacoes", "observações", "descricao", "descricao_completa", "description"],
    }
    for campo, chaves in mapa.items():
        for chave in chaves:
            if chave in item and item[chave]:
                out[campo] = str(item[chave]).strip()
                break
    return out


def ja_existe(ws, item: dict) -> bool:
    """Evita duplicatas: mesmo corretor + bairro + preço."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if (str(row[2] or "").lower() == item["corretor"].lower() and
                str(row[5] or "").lower() == item["bairro"].lower() and
                str(row[10] or "") == item["preco"]):
            return True
    return False


def inserir(item: dict):
    if not PLANILHA.exists():
        print(f"Planilha não encontrada: {PLANILHA}")
        sys.exit(1)

    wb = load_workbook(PLANILHA)
    ws = wb["Imóveis"]

    item = normalizar(item)

    if ja_existe(ws, item):
        print(f"⚠️  Duplicata ignorada: {item['corretor']} / {item['bairro']} / R${item['preco']}")
        return

    next_row = ws.max_row + 1
    fill = ALT_FILL if next_row % 2 == 0 else WHITE_FILL

    valores = [item[k] for k in COL_ORDER]
    for col_idx, val in enumerate(valores, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.fill   = fill
        cell.border = BORDER
        cell.font   = Font(name="Arial", size=10)
        cell.alignment = WRAP if col_idx in (6, 12) else CENTER

    wb.save(PLANILHA)
    print(f"✅ Inserido: {item['corretor']} | {item['bairro']} | R$ {item['preco']}")
    _regenerar_site()


def processar_arquivo(path: str):
    with open(path, "r", encoding="utf-8") as f:
        dados = json.load(f)
    if isinstance(dados, list):
        for item in dados:
            inserir(item)
    else:
        inserir(dados)


def _regenerar_site():
    import subprocess, sys
    site_script = Path(__file__).parent / "gerar_site.py"
    if site_script.exists():
        subprocess.run([sys.executable, str(site_script)], check=False)


def main():
    parser = argparse.ArgumentParser(description="Insere imóveis na planilha a partir do bot WA")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dados", help="JSON do imóvel como string")
    group.add_argument("--arquivo", help="Arquivo JSON com um ou vários imóveis")
    args = parser.parse_args()

    if args.dados:
        inserir(json.loads(args.dados))
    else:
        processar_arquivo(args.arquivo)


if __name__ == "__main__":
    main()
