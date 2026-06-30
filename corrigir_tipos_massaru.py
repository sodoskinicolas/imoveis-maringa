#!/usr/bin/env python3
"""
corrigir_tipos_massaru.py
Visita cada página individual dos imóveis do Massaru no xlsx,
extrai o tipo correto do og:title e corrige a coluna "Tipo".

Uso:
  python3 corrigir_tipos_massaru.py
  python3 corrigir_tipos_massaru.py --dry-run
"""

import re
import sys
import time
import argparse
from pathlib import Path

import requests
import openpyxl

BASE_DIR = Path(__file__).parent
PLANILHA = BASE_DIR / "Imoveis_Grupos.xlsx"

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9",
})

def infer_tipo(s):
    s = (s or "").lower()
    if "sobrado"  in s:                                      return "Sobrado"
    if "kitnet"   in s or "studio" in s or "flat" in s:     return "Kitnet"
    if "cobertura" in s:                                     return "Apartamento"
    if "apart"    in s or "apto"   in s:                    return "Apartamento"
    if "chácara"  in s or "chacara" in s:                   return "Chácara"
    if "sítio"    in s or "sitio"  in s:                    return "Sítio"
    if "galpão"   in s or "galpao" in s:                    return "Galpão"
    if "terreno"  in s or "lote"   in s:                    return "Terreno"
    if "sala"     in s or "comercial" in s or "loja" in s:  return "Sala Comercial"
    if "casa"     in s:                                      return "Casa"
    return None  # desconhecido — não alterar

def get_tipo_from_page(url):
    """Busca og:title da página individual e retorna tipo inferido."""
    try:
        r = SESSION.get(url, timeout=20)
        r.raise_for_status()
        # og:title — forma mais confiável
        m = re.search(r'og:title["\s]+content="([^"]+)"', r.text, re.I)
        if not m:
            m = re.search(r'<title>([^<]+)</title>', r.text, re.I)
        if m:
            return infer_tipo(m.group(1))
    except Exception as e:
        print(f"    ⚠️  Erro ao buscar {url}: {e}")
    return None

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not PLANILHA.exists():
        print(f"❌ Planilha não encontrada: {PLANILHA}")
        sys.exit(1)

    wb = openpyxl.load_workbook(PLANILHA)
    ws = wb["Imóveis"]

    # Descobrir índices das colunas
    headers = {cell.value: cell.column for cell in ws[1]}
    col_grupo  = headers.get("Grupo")
    col_tipo   = headers.get("Tipo")
    col_obs    = headers.get("Observações")

    if not all([col_grupo, col_tipo, col_obs]):
        print(f"❌ Colunas não encontradas. Headers: {list(headers.keys())}")
        sys.exit(1)

    total = 0
    corrigidos = 0

    for row in ws.iter_rows(min_row=2):
        grupo = row[col_grupo - 1].value or ""
        tipo_atual = row[col_tipo - 1].value or ""
        obs = row[col_obs - 1].value or ""

        if "massaru" not in grupo.lower():
            continue

        # Extrair link da obs (formato: "Ref: XXXXX | https://...")
        link_m = re.search(r'https?://[^\s|]+', obs)
        if not link_m:
            continue
        link = link_m.group(0).strip()

        total += 1
        novo_tipo = get_tipo_from_page(link)

        if not novo_tipo:
            print(f"  ⚠️  Sem tipo para {link}")
            continue

        if novo_tipo == tipo_atual:
            print(f"  ✓  OK ({tipo_atual}) — {link}")
            continue

        print(f"  🔄  {tipo_atual} → {novo_tipo}  |  {link}")
        if not args.dry_run:
            row[col_tipo - 1].value = novo_tipo
        corrigidos += 1

        time.sleep(0.5)  # ser gentil com o servidor

    if not args.dry_run and corrigidos > 0:
        wb.save(PLANILHA)
        print(f"\n✅ {corrigidos}/{total} imóveis Massaru corrigidos em {PLANILHA.name}")
    elif args.dry_run:
        print(f"\n[DRY-RUN] {corrigidos}/{total} seriam corrigidos")
    else:
        print(f"\n✅ Tudo já estava correto ({total} imóveis verificados)")

if __name__ == "__main__":
    main()
