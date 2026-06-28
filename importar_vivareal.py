#!/usr/bin/env python3
"""
importar_vivareal.py
Lê vivareal_maringa.json (baixado pelo browser) e salva em:
  - VivaReal_Imoveis.xlsx  (planilha individual)
  - Imoveis_Grupos.xlsx    (planilha padrão, adiciona coluna Data Publicação)

Uso:
    1. Abra a aba do VivaReal no Chrome
    2. Abra o console (F12 → Console)  
    3. Cole e execute:
       window.vrData2 é o array com os dados já coletados.
       Ou baixe novamente: clique em "Salvar JSON" no console aberto no site.
    4. Mova vivareal_maringa.json para esta pasta
    5. Execute: python3 importar_vivareal.py
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale: pip install openpyxl")
    sys.exit(1)

PASTA       = Path(__file__).parent
JSON_FILE   = PASTA / "vivareal_maringa.json"
VR_XLSX     = PASTA / "VivaReal_Imoveis.xlsx"
PLANILHA    = PASTA / "Imoveis_Grupos.xlsx"
HOJE        = datetime.now().strftime("%Y-%m-%d")

TIPO_MAP = {
    'casa':'Casa','apartamento':'Apartamento','lote':'Terreno','terreno':'Terreno',
    'sobrado':'Sobrado','cobertura':'Cobertura','sala':'Sala Comercial',
    'studio':'Studio','kitnet':'Kitnet','chacara':'Chácara/Sítio',
    'galpao':'Galpão','imovel':'Imóvel Comercial','ponto':'Ponto Comercial',
    'predio':'Prédio Comercial','flat':'Studio','duplex':'Sobrado',
}
# Palavras de bairro que aparecem no início do slug → Terreno
BAIRRO_PREFIXOS = {
    'jardim','parque','residencial','vila','loteamento','conjunto','bom','giardino',
    'praca','condominio','sitio','fazenda','zona','setor','district',
}

def limpar_tipo(t, link=''):
    if not t: return 'Imóvel'
    t = t.strip()
    tl = t.lower()
    if tl in TIPO_MAP:
        return TIPO_MAP[tl]
    # Slug começou pelo nome do bairro → Terreno
    slug = link.split('/imovel/')[-1].rstrip('/') if link else ''
    primeiro = slug.split('-')[0].lower() if slug else ''
    if primeiro in BAIRRO_PREFIXOS:
        return 'Terreno'
    # Tipo com inicial minúscula = parse ruim do slug → Terreno
    if t[0].islower():
        return 'Terreno'
    return t

# ─── Ler JSON ────────────────────────────────────────────────────────────────
if not JSON_FILE.exists():
    print(f"❌ Arquivo não encontrado: {JSON_FILE}")
    print()
    print("Para gerar o arquivo:")
    print("  1. Abra o site VivaReal no Chrome")
    print("  2. No console do browser (F12), execute:")
    print("""
const json = JSON.stringify(window.vrData2);
const b = new Blob([json],{type:'application/json'});
const a = document.createElement('a');
a.href = URL.createObjectURL(b);
a.download = 'vivareal_maringa.json';
a.click();
""")
    print("  3. Mova o arquivo baixado para esta pasta e rode novamente.")
    sys.exit(1)

with open(JSON_FILE, encoding='utf-8') as f:
    raw = json.load(f)

print(f"📂 {len(raw)} imóveis lidos de {JSON_FILE.name}")

# Normalizar campos
listings = []
for r in raw:
    if not r.get('id'): continue
    tipo = limpar_tipo(r.get('tipo') or r.get('t') or '', r.get('link') or r.get('lk') or '')
    listings.append({
        'id':        str(r.get('id') or r.get('id','')),
        'tipo':      tipo,
        'bairro':    r.get('bairro') or r.get('b') or '',
        'rua':       r.get('rua') or r.get('r') or '',
        'area':      r.get('area') or r.get('a'),
        'quartos':   r.get('quartos') or r.get('q'),
        'banheiros': r.get('banheiros') or r.get('bn'),
        'vagas':     r.get('vagas') or r.get('v'),
        'preco':     r.get('preco') or r.get('p'),
        'corretor':  r.get('corretor') or '',
        'dataPub':   r.get('dataPub') or r.get('data_publicacao') or '',
        'link':      r.get('link') or r.get('lk') or '',
    })

print(f"✅ {len(listings)} imóveis válidos")

# ─── Criar VivaReal_Imoveis.xlsx ────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "VivaReal Maringá"

HEADER  = ["ID VivaReal","Data Publicação","Data Captura","Tipo","Bairro",
           "Endereço","Área (m²)","Quartos","Banheiros","Vagas","Preço (R$)","Corretor","Link"]
WIDTHS  = [16,16,14,16,28,36,10,8,10,8,14,30,60]
H_FILL  = PatternFill("solid", fgColor="1F4E79")
H_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin    = Side(style='thin', color='DDDDDD')
BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)

for ci, (h, w) in enumerate(zip(HEADER, WIDTHS), 1):
    c = ws.cell(1, ci, h)
    c.font      = H_FONT
    c.fill      = H_FILL
    c.alignment = H_ALIGN
    c.border    = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

ALT_FILL = PatternFill("solid", fgColor="F5F8FF")

for ri, im in enumerate(listings, 2):
    row = [
        im['id'], im['dataPub'], HOJE, im['tipo'],
        im['bairro'], im['rua'],
        im['area'], im['quartos'], im['banheiros'], im['vagas'],
        im['preco'], im['corretor'], im['link'],
    ]
    for ci, val in enumerate(row, 1):
        c = ws.cell(ri, ci, val)
        c.border = BORDER
        if ri % 2 == 0:
            c.fill = ALT_FILL

wb.save(VR_XLSX)
print(f"💾 Salvo: {VR_XLSX.name}  ({len(listings)} linhas)")

# ─── Atualizar Imoveis_Grupos.xlsx ──────────────────────────────────────────
if not PLANILHA.exists():
    print(f"⚠ Planilha padrão não encontrada, pulando: {PLANILHA}")
    sys.exit(0)

wb2 = load_workbook(PLANILHA)
ws2 = wb2["Imóveis"]

# Ensure "Data Publicação" column exists
headers = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
if "Data Publicação" not in headers:
    col_dp = ws2.max_column + 1
    c = ws2.cell(1, col_dp, "Data Publicação")
    c.font  = Font(bold=True, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="1F4E79")
    ws2.column_dimensions[get_column_letter(col_dp)].width = 16
else:
    col_dp = headers.index("Data Publicação") + 1

# Collect existing VR IDs
existing_ids = set()
for row in ws2.iter_rows(min_row=2, values_only=True):
    obs = str(row[11] or '')
    if 'id:' in obs:
        try: existing_ids.add(obs.split('id:')[1].strip())
        except: pass

inseridos = 0
for im in listings:
    if im['id'] in existing_ids: continue
    bairro_end = f"{im['bairro']} · {im['rua']}".strip(' ·') if im['rua'] else im['bairro']
    ws2.append([
        HOJE,                                        # Data Captura
        "vivareal.com.br",                           # Grupo
        im['corretor'] or "VivaReal",               # Corretor
        "",                                          # Contato WhatsApp
        im['tipo'],                                  # Tipo
        bairro_end,                                  # Bairro / Endereço
        im['area'],                                  # Área (m²)
        im['quartos'],                               # Quartos
        None,                                        # Suítes
        im['vagas'],                                 # Vagas
        im['preco'],                                 # Preço (R$)
        f"VivaReal | id:{im['id']}",                 # Observações
        "Venda",                                     # Status
        im['dataPub'],                               # Data Publicação (col 14)
    ])
    existing_ids.add(im['id'])
    inseridos += 1

wb2.save(PLANILHA)
print(f"💾 Atualizado: {PLANILHA.name}  (+{inseridos} novos imóveis VivaReal)")
print()
print("✅ Tudo pronto! Rode: python3 gerar_site.py")
