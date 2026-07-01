"""
Validador de WhatsApp - Ed. São Gabriel
Usa a API 2chat para checar quais números têm WhatsApp.
Execute: python3 validar_saogabriel.py
"""

import re
import requests
import time
import openpyxl
from openpyxl.styles import PatternFill, Font

API_KEY = "UAK3057d948-badf-43be-ada2-d94f2ac80814"
MEU_NUMERO = "+5544991529335"
BASE_URL = "https://api.p.2chat.io/open/whatsapp/check-number"

PROPRIETARIOS = [
    {"apt": "Apartamento 202",  "owner": "Fernando Henrique Da Cruz",            "numbers": []},
    {"apt": "Apartamento 208",  "owner": "Fernando Henrique Da Cruz",            "numbers": []},
    {"apt": "Apartamento 301",  "owner": "Fernando Henrique Da Cruz",            "numbers": []},
    {"apt": "Apartamento 303",  "owner": "Fernando Henrique Da Cruz",            "numbers": []},
    {"apt": "Apartamento 307",  "owner": "Fernando Henrique Da Cruz",            "numbers": []},
    {"apt": "Apartamento 404",  "owner": "Denisley Vicentino",                   "numbers": ["+5569981273189", "+5569981082309", "+5544997565322"]},
    {"apt": "Apartamento 405",  "owner": "Cesar Axiel Sgobero",                  "numbers": ["+5544988032629", "+5544988034098", "+5544991216698"]},
    {"apt": "Apartamento 406",  "owner": "Denisley Vicentino",                   "numbers": ["+5569981273189", "+5569981082309", "+5544997565322"]},
    {"apt": "Apartamento 503",  "owner": "Denisley Vicentino",                   "numbers": ["+5569981273189", "+5569981082309", "+5544997565322"]},
    {"apt": "Apartamento 508",  "owner": "Denisley Vicentino",                   "numbers": ["+5569981273189", "+5569981082309", "+5544997565322"]},
    {"apt": "Apartamento 603",  "owner": "Denisley Vicentino",                   "numbers": ["+5569981273189", "+5569981082309", "+5544997565322"]},
    {"apt": "Apartamento 607",  "owner": "Wilson Senhorinho",                    "numbers": ["+5565981162055", "+5544991198855", "+5544991190606"]},
    {"apt": "Apartamento 608",  "owner": "Gabriel Fernandes Manduca",            "numbers": ["+5543999838587"]},
    {"apt": "Apartamento 701",  "owner": "Gabriel Fernandes Manduca",            "numbers": ["+5543999838587"]},
    {"apt": "Apartamento 702",  "owner": "Antonio Carlos De Paula",              "numbers": ["+5544999728261", "+5544999429663"]},
    {"apt": "Apartamento 703",  "owner": "Katia Regina Aguiar",                  "numbers": ["+5544999730601", "+5544999849260"]},
    {"apt": "Apartamento 705",  "owner": "Fouad Hassan Paracat",                 "numbers": ["+5543996623362"]},
    {"apt": "Apartamento 708",  "owner": "Fouad Hassan Paracat",                 "numbers": ["+5543996623362"]},
    {"apt": "Apartamento 805",  "owner": "Julio Cesar Calixto",                  "numbers": ["+5548996320100", "+5544997078989"]},
    {"apt": "Apartamento 1108", "owner": "Evanilson Picelli",                    "numbers": ["+5544999287645", "+5544999730566"]},
    {"apt": "Apartamento 1204", "owner": "Evanilson Picelli",                    "numbers": ["+5544999287645", "+5544999730566"]},
    {"apt": "Apartamento 1205", "owner": "Ailton De Souza",                      "numbers": ["+5544991014402", "+5544998711209"]},
    {"apt": "Apartamento 1206", "owner": "Denisley Vicentino",                   "numbers": ["+5569981273189", "+5569981082309", "+5544997565322"]},
    {"apt": "Apartamento 1207", "owner": "Jaime Luiz Carlesso",                  "numbers": ["+5544991162262", "+5544997357726"]},
    {"apt": "Apartamento 1208", "owner": "Silvio Koitiro Siraichi",              "numbers": ["+5544999731470"]},
    {"apt": "Apartamento 1302", "owner": "Jiro Numoto",                          "numbers": ["+5544999675643", "+5541998419285"]},
    {"apt": "Apartamento 1303", "owner": "Ericson Fabiano Marcossi",             "numbers": ["+5544988144926", "+5544991059679", "+5544991685029", "+5544984233484"]},
    {"apt": "Apartamento 1304", "owner": "Jiro Numoto",                          "numbers": ["+5544999675643", "+5541998419285"]},
    {"apt": "Apartamento 1305", "owner": "Dalila Maria Cristina De Souza Paz",   "numbers": ["+5544998302324", "+5544997714929", "+5544984495119"]},
    {"apt": "Apartamento 1306", "owner": "Renato Kraviski",                      "numbers": ["+5541996022357", "+5541992377916", "+5541999977613", "+5541991290272"]},
    {"apt": "Apartamento 1307", "owner": "Renato Kraviski",                      "numbers": ["+5541996022357", "+5541992377916", "+5541999977613", "+5541991290272"]},
    {"apt": "Apartamento 1308", "owner": "Renato Kraviski",                      "numbers": ["+5541996022357", "+5541992377916", "+5541999977613", "+5541991290272"]},
    {"apt": "Apartamento 1401", "owner": "Renato Kraviski",                      "numbers": ["+5541996022357", "+5541992377916", "+5541999977613", "+5541991290272"]},
    {"apt": "Apartamento 1402", "owner": "Renato Kraviski",                      "numbers": ["+5541996022357", "+5541992377916", "+5541999977613", "+5541991290272"]},
    {"apt": "Apartamento 1404", "owner": "Armid Wesley Formigoni Dias",          "numbers": ["+5547991851609", "+5532998219939", "+5544998212125", "+5544999684348"]},
    {"apt": "Apartamento 1405", "owner": "Raissa Cortez Sanchez",                "numbers": ["+5544984380252", "+5544999314414"]},
    {"apt": "Apartamento 1406", "owner": "Maria Iolanda Dos Santos Rocha",       "numbers": ["+5541991151451"]},
    {"apt": "Apartamento 1407", "owner": "Claudemir Da Silva Rosa",              "numbers": ["+5544999734143", "+5544997070415"]},
    {"apt": "Apartamento 1408", "owner": "Claudemir Da Silva Rosa",              "numbers": ["+5544999734143", "+5544997070415"]},
    {"apt": "Apartamento 1504", "owner": "Claudemir Da Silva Rosa",              "numbers": ["+5544999734143", "+5544997070415"]},
    {"apt": "Apartamento 1506", "owner": "Charles Henrique Mendes",              "numbers": ["+5544991184738", "+5545999909090", "+5545998166767"]},
    {"apt": "Apartamento 1602", "owner": "Jose Fernando Fedossi",                "numbers": ["+5517991426470", "+5517981846000", "+5517997177723"]},
    {"apt": "Apartamento 1607", "owner": "Jose Fernando Fedossi",                "numbers": ["+5517991426470", "+5517981846000", "+5517997177723"]},
    {"apt": "Apartamento 1702", "owner": "Arnoldo Alencar Arduim",               "numbers": ["+5544999735791", "+5544988365913"]},
    {"apt": "Apartamento 1705", "owner": "Arnoldo Alencar Arduim",               "numbers": ["+5544999735791", "+5544988365913"]},
    {"apt": "Apartamento 1707", "owner": "Cidinei Aparecido Vaz",                "numbers": ["+5544999309339", "+5544999806261", "+5544991753850"]},
    {"apt": "Apartamento 1801", "owner": "Cesar Axiel Sgobero",                  "numbers": ["+5544988032629", "+5544988034098", "+5544991216698"]},
    {"apt": "Apartamento 1802", "owner": "Cesar Axiel Sgobero",                  "numbers": ["+5544988032629", "+5544988034098", "+5544991216698"]},
    {"apt": "Apartamento 1804", "owner": "Cesar Axiel Sgobero",                  "numbers": ["+5544988032629", "+5544988034098", "+5544991216698"]},
]

def check_number(number):
    url = f"{BASE_URL}/{MEU_NUMERO}/{number}"
    try:
        resp = requests.get(url, headers={"X-User-API-Key": API_KEY}, timeout=15)
        data = resp.json()
        return data.get("on_whatsapp", False)
    except Exception as e:
        print(f"  ERRO: {e}")
        return None

def main():
    results = []
    total = sum(len(p["numbers"]) for p in PROPRIETARIOS)
    count = 0

    print(f"Validando {total} números de {len(PROPRIETARIOS)} proprietários...\n")

    for prop in PROPRIETARIOS:
        if not prop["numbers"]:
            print(f"  Apto {prop['apt']} | {prop['owner']} | Sem número cadastrado")
            results.append({"apt": prop["apt"], "owner": prop["owner"],
                            "has_whatsapp": False, "wa_numbers": [], "all_numbers": []})
            continue

        number_results = []
        for num in prop["numbers"]:
            count += 1
            on_wa = check_number(num)
            status = "✅ WhatsApp" if on_wa else ("❌ Sem WA" if on_wa is False else "⚠️ Erro")
            print(f"[{count}/{total}] {prop['apt']} | {num} | {status}")
            number_results.append({"number": num, "on_whatsapp": on_wa})
            time.sleep(0.5)

        wa_numbers = [n["number"] for n in number_results if n["on_whatsapp"]]
        results.append({
            "apt": prop["apt"],
            "owner": prop["owner"],
            "has_whatsapp": len(wa_numbers) > 0,
            "wa_numbers": wa_numbers,
            "all_numbers": number_results,
        })

    # Descobrir máximo de números com WA para colunas dinâmicas
    max_wa = max((len(r["wa_numbers"]) for r in results), default=1)

    # Gerar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validação WhatsApp"

    num_cols = [f"WhatsApp {i+1}" for i in range(max_wa)]
    headers = ["Apartamento", "Proprietário", "Tem WhatsApp?"] + num_cols
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFCDD2")

    for r in results:
        wa_cols = r["wa_numbers"] + [""] * (max_wa - len(r["wa_numbers"]))
        row = [r["apt"], r["owner"], "SIM" if r["has_whatsapp"] else "NÃO"] + wa_cols
        ws.append(row)
        fill = green if r["has_whatsapp"] else red
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    for i in range(max_wa):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = 22

    output = "SaoGabriel_WhatsApp_Validado.xlsx"
    wb.save(output)

    total_wa = sum(1 for r in results if r["has_whatsapp"])
    print(f"\n✅ Concluído! {total_wa}/{len(results)} proprietários com WhatsApp.")
    print(f"📊 Planilha salva: {output}")

if __name__ == "__main__":
    main()
